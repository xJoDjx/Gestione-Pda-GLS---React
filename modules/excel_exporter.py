"""
excel_exporter.py - Esportazione Excel formattata (stile originale)
"""
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers as openpyxl_numbers)
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import date
from modules.database import load_padroncini, load_conteggi, get_conteggio

MONTHS_IT = [
    '', 'GENNAIO', 'FEBBRAIO', 'MARZO', 'APRILE', 'MAGGIO', 'GIUGNO',
    'LUGLIO', 'AGOSTO', 'SETTEMBRE', 'OTTOBRE', 'NOVEMBRE', 'DICEMBRE'
]

# ─── Colors ───────────────────────────────────────────────────────────────────
C_HEADER_BG  = "1F3864"   # dark blue
C_HEADER_FG  = "FFFFFF"
C_SECTION_BG = "2F5496"   # medium blue
C_SECTION_FG = "FFFFFF"
C_LABEL_BG   = "D6E4F7"   # light blue
C_VALUE_BG   = "FFFFFF"
C_TOTAL_BG   = "FFE699"   # yellow for totals
C_RED_BG     = "FCE4D6"   # light red for debits
C_GREEN_BG   = "E2EFDA"   # light green for credits
C_ALT_ROW    = "F2F2F2"
C_ATTIVO     = "C6EFCE"
C_SCADUTO    = "FFC7CE"
C_DISMESSO   = "FFEB9C"

BORDER_THIN = Border(
    left=Side(style='thin', color='BFBFBF'),
    right=Side(style='thin', color='BFBFBF'),
    top=Side(style='thin', color='BFBFBF'),
    bottom=Side(style='thin', color='BFBFBF')
)

BORDER_MEDIUM = Border(
    left=Side(style='medium', color='595959'),
    right=Side(style='medium', color='595959'),
    top=Side(style='medium', color='595959'),
    bottom=Side(style='medium', color='595959')
)


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, size=10, color="000000", italic=False):
    return Font(name='Arial', bold=bold, size=size, color=color, italic=italic)


def _align(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _set_cell(ws, row, col, value, font=None, fill=None, alignment=None, border=None, number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:       cell.font = font
    if fill:       cell.fill = fill
    if alignment:  cell.alignment = alignment
    if border:     cell.border = border
    if number_format: cell.number_format = number_format
    return cell


def _merge(ws, r1, c1, r2, c2, value='', font=None, fill=None, alignment=None, border=None):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(row=r1, column=c1, value=value)
    if font:       cell.font = font
    if fill:       cell.fill = fill
    if alignment:  cell.alignment = alignment
    return cell


# ─── RIEPILOGO sheet ──────────────────────────────────────────────────────────

def _build_riepilogo(ws, padroncini, conteggi, mese, anno):
    ws.sheet_view.showGridLines = False

    # Column widths
    col_widths = [3, 30, 16, 16, 16, 12, 12, 12, 12, 12, 12, 16, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title
    _merge(ws, 1, 1, 1, 13,
           f"RIEPILOGO MENSILE GLS  –  {mese} {anno}",
           font=_font(bold=True, size=14, color=C_HEADER_FG),
           fill=_fill(C_HEADER_BG),
           alignment=_align('center'))
    ws.row_dimensions[1].height = 28

    # Sub header
    headers = ['', 'FORNITORE', 'TOT FATTURATO', 'TOT ADDEBITO', 'BONIFICO',
               'STATO', 'DURC', 'N.PALMARI', 'N.MEZZI', 'DISTR', 'PDF', 'FATT.RIC', 'SCAD.']
    for j, h in enumerate(headers, 1):
        _set_cell(ws, 2, j, h,
                  font=_font(bold=True, size=9, color=C_HEADER_FG),
                  fill=_fill(C_SECTION_BG),
                  alignment=_align('center'),
                  border=BORDER_THIN)
    ws.row_dimensions[2].height = 20

    # Build lookup
    cont_map = {}
    for c in conteggi:
        if c.get('mese') == mese and c.get('anno') == anno:
            cont_map[c['padroncino_id']] = c

    totale_fatt = totale_add = totale_bon = 0
    row = 3
    for idx, p in enumerate(padroncini):
        c = cont_map.get(p['id'], {})
        fatt = c.get('totale_imponibile', 0)
        add  = c.get('totale_addebiti', 0)
        bon  = c.get('totale_da_bonificare', 0)

        stato = p.get('stato', 'ATTIVO')
        durc  = p.get('durc_stato', '')
        bg_row = C_ALT_ROW if idx % 2 else C_VALUE_BG

        palm_list = p.get('palmari', [])
        mez_list  = p.get('mezzi', [])
        n_palm = len(palm_list) if isinstance(palm_list, list) else 0
        n_mez  = len(mez_list)  if isinstance(mez_list, list) else 0
        vals = [idx+1, p['nome'], fatt, add, bon, stato, durc,
                n_palm, n_mez,
                '✓' if c.get('distrib_inviata') else '',
                '✓' if c.get('pdf_addeb') else '',
                '✓' if c.get('fattura_ricevuta') else '',
                '✓' if c.get('caricata_scadenziario') else '']

        for j, v in enumerate(vals, 1):
            fmt = '#,##0.00' if j in (3, 4, 5) else None
            col_fill = _fill(bg_row)
            if j == 6:  # stato
                col_fill = _fill(C_ATTIVO if stato == 'ATTIVO' else
                                 C_DISMESSO if 'DISMESSO' in stato else C_SCADUTO)
            if j == 7:  # durc
                col_fill = _fill(C_ATTIVO if durc == 'VALIDO' else C_SCADUTO)
            _set_cell(ws, row, j, v,
                      font=_font(size=9),
                      fill=col_fill,
                      alignment=_align('center' if j != 2 else 'left'),
                      border=BORDER_THIN,
                      number_format=fmt)
        totale_fatt += fatt
        totale_add  += add
        totale_bon  += bon
        row += 1

    # Totals row
    _merge(ws, row, 1, row, 2, 'TOTALE',
           font=_font(bold=True, size=10),
           fill=_fill(C_TOTAL_BG),
           alignment=_align('center'))
    for col, val in [(3, totale_fatt), (4, totale_add), (5, totale_bon)]:
        _set_cell(ws, row, col, val,
                  font=_font(bold=True, size=10),
                  fill=_fill(C_TOTAL_BG),
                  alignment=_align('center'),
                  border=BORDER_THIN,
                  number_format='#,##0.00')
    ws.row_dimensions[row].height = 18


# ─── Single padroncino sheet ──────────────────────────────────────────────────

def _build_padroncino_sheet(ws, padroncino, conteggio, mese, anno):
    ws.sheet_view.showGridLines = False

    col_w = [2, 32, 14, 22, 14, 14, 2, 28, 14, 14, 2, 30]
    for i, w in enumerate(col_w, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    nome = padroncino.get('nome', '')
    codice = padroncino.get('codice', '')

    # ── INTESTAZIONE ──────────────────────────────────────────────
    _merge(ws, 1, 1, 2, 12, f"{nome}  ({codice})  –  {mese} {anno}",
           font=_font(bold=True, size=13, color=C_HEADER_FG),
           fill=_fill(C_HEADER_BG),
           alignment=_align('center'))
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 8

    # DURC info
    durc_sc = padroncino.get('durc_scadenza', '')
    durc_st = padroncino.get('durc_stato', '')
    durc_fill = _fill(C_ATTIVO if durc_st == 'VALIDO' else C_SCADUTO)
    _set_cell(ws, 3, 2, f"DURC SCADENZA: {durc_sc}",
              font=_font(bold=True, size=9), fill=_fill(C_LABEL_BG),
              alignment=_align('left'), border=BORDER_THIN)
    _set_cell(ws, 3, 3, durc_st,
              font=_font(bold=True, size=9), fill=durc_fill,
              alignment=_align('center'), border=BORDER_THIN)

    # Note varie
    note = conteggio.get('note_varie', '') if conteggio else ''
    _merge(ws, 3, 8, 3, 12, f"NOTE: {note}",
           font=_font(italic=True, size=8, color='595959'),
           fill=_fill('FFFBE6'),
           alignment=_align('left', wrap=True))

    r = 5  # current row

    # ── SEZIONE FATTURATO ─────────────────────────────────────────
    r = _section_header(ws, r, 12, "FATTURATO SPEDIZIONI + RITIRI")
    r = _label_value_row(ws, r, "Fisso Mensile", conteggio, 'fisso_mensile', bg=C_LABEL_BG)
    r = _label_value_row(ws, r, "Totale Spedizioni (Proforma)", conteggio, 'totale_spedizioni')
    r = _label_value_row(ws, r, "Totale Ritiri", conteggio, 'totale_ritiri')
    r = _label_value_row(ws, r, "Totale Ritiri Zona Fissi", conteggio, 'totale_ritiri_fissi')
    r = _label_value_row(ws, r, "Consegne Doppie", conteggio, 'consegne_doppie', negative=True)
    r = _label_value_row(ws, r, "Sforamento Rientri", conteggio, 'sforamento_rientri', negative=True)
    r = _label_value_row(ws, r, "Compensazioni su Imponibile", conteggio, 'compensazioni_imponibile')

    # Extra items
    for item in (conteggio or {}).get('altri_fatturato', []):
        r = _custom_row(ws, r, item.get('descrizione', ''), item.get('importo', 0))

    r += 1
    r = _section_header(ws, r, 5, "TOTALE DOVUTO")
    r = _label_value_row(ws, r, "Totale Imponibile", conteggio, 'totale_imponibile', bold=True, bg=C_TOTAL_BG)
    r = _label_value_row(ws, r, "IVA (22%)", conteggio, 'iva')
    r = _label_value_row(ws, r, "Totale Fattura da Ricevere", conteggio, 'totale_fattura', bold=True, bg=C_TOTAL_BG)
    r += 1

    # ── SEZIONE ADDEBITI ──────────────────────────────────────────
    r = _section_header_right(ws, r, 12, "COMPENSAZIONI E ADDEBITI SU DISTRIBUZIONE")
    comp_row = r
    r += 1

    # Left: addebiti details
    r_left = comp_row + 1
    _set_cell(ws, r_left, 2, "ADDEBITI PALMARI",
              font=_font(bold=True, size=9), fill=_fill(C_LABEL_BG),
              alignment=_align('left'), border=BORDER_THIN)
    r_left += 1
    n_palm = (conteggio or {}).get('n_palmari', 0)
    add_palm = (conteggio or {}).get('addebiti_palmari', 0)
    _set_cell(ws, r_left, 2, f"N. Palmari: {n_palm}",
              font=_font(size=9), alignment=_align(), border=BORDER_THIN)
    _set_cell(ws, r_left, 3, add_palm,
              font=_font(size=9), alignment=_align('center'),
              border=BORDER_THIN, number_format='#,##0.00')
    r_left += 1

    _set_cell(ws, r_left, 2, "ADDEBITI MEZZI",
              font=_font(bold=True, size=9), fill=_fill(C_LABEL_BG),
              alignment=_align(), border=BORDER_THIN)
    r_left += 1
    for mezzo in (conteggio or {}).get('dettagli_mezzi', []):
        _set_cell(ws, r_left, 2, mezzo.get('targa', ''),
                  font=_font(size=9), alignment=_align(), border=BORDER_THIN)
        _set_cell(ws, r_left, 3, mezzo.get('importo', 0),
                  font=_font(size=9), alignment=_align('center'),
                  border=BORDER_THIN, number_format='#,##0.00')
        _set_cell(ws, r_left, 4, mezzo.get('tipologia', ''),
                  font=_font(size=8, italic=True), alignment=_align(), border=BORDER_THIN)
        r_left += 1

    for item in (conteggio or {}).get('altri_addebiti', []):
        _set_cell(ws, r_left, 2, item.get('descrizione', ''),
                  font=_font(size=9), alignment=_align(), border=BORDER_THIN)
        _set_cell(ws, r_left, 3, item.get('importo', 0),
                  font=_font(size=9), fill=_fill(C_RED_BG), alignment=_align('center'),
                  border=BORDER_THIN, number_format='#,##0.00')
        r_left += 1

    # Right: compensazioni, fatture fine mese, cassa prima nota
    r_right = comp_row + 1
    _set_cell(ws, r_right, 8, "ACCREDITI E STORNI",
              font=_font(bold=True, size=9), fill=_fill(C_SECTION_BG),
              alignment=_align('center'), border=BORDER_THIN)
    _merge(ws, r_right, 9, r_right, 10, 'VAL',
           font=_font(bold=True, size=9, color=C_HEADER_FG),
           fill=_fill(C_SECTION_BG), alignment=_align('center'))
    r_right += 1

    for ff in (conteggio or {}).get('fatture_fine_mese', []):
        _set_cell(ws, r_right, 8, ff.get('descrizione', ''),
                  font=_font(size=8), fill=_fill(C_GREEN_BG),
                  alignment=_align('left', wrap=True), border=BORDER_THIN)
        _set_cell(ws, r_right, 9, ff.get('importo', 0),
                  font=_font(size=9, bold=True), fill=_fill(C_GREEN_BG),
                  alignment=_align('center'), border=BORDER_THIN, number_format='#,##0.00')
        r_right += 1

    r_right += 1
    _set_cell(ws, r_right, 8, "CASSA PRIMA NOTA",
              font=_font(bold=True, size=9), fill=_fill(C_LABEL_BG),
              alignment=_align(), border=BORDER_THIN)
    r_right += 1
    for item in (conteggio or {}).get('cassa_prima_nota', []):
        _set_cell(ws, r_right, 8, item.get('descrizione', ''),
                  font=_font(size=8), alignment=_align('left', wrap=True), border=BORDER_THIN)
        _set_cell(ws, r_right, 9, item.get('importo', 0),
                  font=_font(size=9), alignment=_align('center'),
                  border=BORDER_THIN, number_format='#,##0.00')
        r_right += 1

    # Final totals (right side)
    max_r = max(r_left, r_right) + 1
    _set_cell(ws, max_r, 8, "TOTALE ADDEBITI DA INSERIRE IN FATTURA",
              font=_font(bold=True, size=9), fill=_fill(C_TOTAL_BG),
              alignment=_align(), border=BORDER_MEDIUM)
    _set_cell(ws, max_r, 9, (conteggio or {}).get('totale_addebiti', 0),
              font=_font(bold=True, size=10), fill=_fill(C_TOTAL_BG),
              alignment=_align('center'), border=BORDER_MEDIUM, number_format='#,##0.00')
    max_r += 1

    bon = (conteggio or {}).get('totale_da_bonificare', 0)
    bon_fill = C_GREEN_BG if bon >= 0 else C_RED_BG
    _set_cell(ws, max_r, 8, "TOTALE DA BONIFICARE",
              font=_font(bold=True, size=11), fill=_fill(bon_fill),
              alignment=_align(), border=BORDER_MEDIUM)
    _set_cell(ws, max_r, 9, bon,
              font=_font(bold=True, size=12), fill=_fill(bon_fill),
              alignment=_align('center'), border=BORDER_MEDIUM, number_format='#,##0.00')

    # ── CHECKLIST row ─────────────────────────────────────────────
    ck_row = max_r + 2
    checks = [
        ('DISTR. INVIATA', 'distrib_inviata'),
        ('PDF ADDEB.', 'pdf_addeb'),
        ('FATT. RICEV.', 'fattura_ricevuta'),
        ('FATT. TU CREAT.', 'fatt_tu_creata'),
        ('UNIONE PDF', 'unione_pdf'),
        ('SCADENZ.', 'caricata_scadenziario'),
    ]
    _merge(ws, ck_row, 1, ck_row, 12, '',
           fill=_fill(C_SECTION_BG))
    ck_row += 1
    for idx, (label, key) in enumerate(checks):
        col = 2 + idx * 2
        done = (conteggio or {}).get(key, False)
        _set_cell(ws, ck_row, col, label,
                  font=_font(size=8, bold=True),
                  fill=_fill(C_LABEL_BG),
                  alignment=_align('center'), border=BORDER_THIN)
        _set_cell(ws, ck_row, col + 1, '✓' if done else '',
                  font=_font(size=12, bold=True, color='006400' if done else 'CC0000'),
                  alignment=_align('center'), border=BORDER_THIN)


def _section_header(ws, r, cols, title, fg=C_SECTION_FG, bg=C_SECTION_BG):
    _merge(ws, r, 2, r, cols, title,
           font=_font(bold=True, size=10, color=fg),
           fill=_fill(bg),
           alignment=_align('center'))
    ws.row_dimensions[r].height = 18
    return r + 1


def _section_header_right(ws, r, cols, title):
    _merge(ws, r, 8, r, cols, title,
           font=_font(bold=True, size=9, color=C_HEADER_FG),
           fill=_fill(C_SECTION_BG),
           alignment=_align('center'))
    return r + 1


def _label_value_row(ws, r, label, conteggio, key, bold=False, bg=None, negative=False):
    val = (conteggio or {}).get(key, 0)
    row_bg = bg or (C_RED_BG if (negative and val < 0) else C_VALUE_BG)
    _set_cell(ws, r, 2, label,
              font=_font(bold=bold, size=9), fill=_fill(C_LABEL_BG),
              alignment=_align(), border=BORDER_THIN)
    _set_cell(ws, r, 3, val,
              font=_font(bold=bold, size=9), fill=_fill(row_bg),
              alignment=_align('center'), border=BORDER_THIN,
              number_format='#,##0.00')
    ws.row_dimensions[r].height = 16
    return r + 1


def _custom_row(ws, r, label, val):
    _set_cell(ws, r, 2, label,
              font=_font(size=9), fill=_fill(C_LABEL_BG),
              alignment=_align(), border=BORDER_THIN)
    _set_cell(ws, r, 3, val,
              font=_font(size=9), fill=_fill(C_VALUE_BG),
              alignment=_align('center'), border=BORDER_THIN,
              number_format='#,##0.00')
    return r + 1


# ─── Main export function ─────────────────────────────────────────────────────

def export_excel(mese: str, anno: int, output_path: str = None) -> str:
    """Esporta il file Excel mensile completo"""
    padroncini = load_padroncini()
    conteggi   = load_conteggi()

    cont_map = {c['padroncino_id']: c
                for c in conteggi
                if c.get('mese') == mese and c.get('anno') == anno}

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # RIEPILOGO sheet first
    ws_riepilogo = wb.create_sheet("RIEPILOGO")
    _build_riepilogo(ws_riepilogo, padroncini, conteggi, mese, anno)

    # One sheet per padroncino
    for p in padroncini:
        safe_name = p['nome'][:28].strip()
        ws = wb.create_sheet(safe_name)
        c  = cont_map.get(p['id'])
        _build_padroncino_sheet(ws, p, c, mese, anno)

    if not output_path:
        exports_dir = Path(__file__).parent.parent / "exports"
        exports_dir.mkdir(exist_ok=True)
        output_path = str(exports_dir / f"GLS_{mese}_{anno}_Conteggi.xlsx")

    wb.save(output_path)
    print(f"Excel salvato: {output_path}")
    return output_path
